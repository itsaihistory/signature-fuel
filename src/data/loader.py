"""
Reference data loader — reads FBO locations, terminal mappings,
pricing index assignments, and transport rates from Excel/CSV files.
"""

import csv
import json
import logging
from pathlib import Path
from typing import Optional

from src.models.fbo import FBO, Terminal, TransportRate, MarketConfig, Contract
from config.settings import REFERENCE_DIR, STATE_TO_PADD

logger = logging.getLogger(__name__)

# Default paths to reference files (copied into data/reference/)
DOWNLOADS = Path("C:/Users/parke/Downloads")


def _try_float(val, default=None):
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


class ReferenceDataLoader:
    """Loads and parses Signature reference data files."""

    def __init__(self, ref_dir: Optional[Path] = None):
        self.ref_dir = ref_dir or REFERENCE_DIR

    def load_fbos(self, path: Optional[Path] = None) -> list[FBO]:
        """Load FBO locations from the pricing index file (most complete)."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed — trying CSV fallback")
            return self._load_fbos_csv(path)

        xlsx_path = path or self._find_file(
            "signature_fbo_with_pricing_index_v3.xlsx",
            "Signature FBOs.xlsx",
        )
        if not xlsx_path:
            logger.error("No FBO file found")
            return []

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        fbos = []
        for row in rows:
            if not row or not row[0]:
                continue
            code = str(row[0]).strip()
            city = str(row[1]).strip() if row[1] else ""
            state = str(row[2]).strip() if row[2] else ""

            fbo = FBO(
                code=code,
                city=city,
                state=state,
                latitude=_try_float(row[3]) if len(row) > 3 else None,
                longitude=_try_float(row[4]) if len(row) > 4 else None,
                padd=STATE_TO_PADD.get(state),
                index_publisher=str(row[5]).strip() if len(row) > 5 and row[5] else "Platts",
                index_name=str(row[6]).strip() if len(row) > 6 and row[6] else "",
                sample_differential=_try_float(row[7], 0.0) if len(row) > 7 else 0.0,
                region_label=str(row[8]).strip() if len(row) > 8 and row[8] else "",
            )
            fbos.append(fbo)

        wb.close()
        logger.info(f"Loaded {len(fbos)} FBOs from {xlsx_path.name}")
        return fbos

    def _load_fbos_csv(self, path: Optional[Path] = None) -> list[FBO]:
        """Fallback CSV loader for FBO data."""
        csv_path = path or self._find_file("FBOs_with_Pricing_Index_Assignments.csv")
        if not csv_path:
            return []

        fbos = []
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                code = row.get("FBO", "").strip()
                if not code:
                    continue
                state = row.get("State", "").strip()
                fbos.append(FBO(
                    code=code,
                    city=row.get("City", "").strip(),
                    state=state,
                    padd=STATE_TO_PADD.get(state),
                    index_publisher=row.get("Index_Publisher", "Platts").strip(),
                    index_name=row.get("Index_Name", "").strip(),
                    sample_differential=_try_float(row.get("Sample_Differential_$/gal"), 0.0),
                ))

        logger.info(f"Loaded {len(fbos)} FBOs from CSV")
        return fbos

    def load_terminals(self, path: Optional[Path] = None) -> list[Terminal]:
        """Load terminal directory."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed")
            return []

        xlsx_path = path or self._find_file("Terminals.xlsx", "FBO_to_Nearest_Terminals.xlsx")
        if not xlsx_path:
            return []

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # Try "Terminals" sheet first
        ws = None
        for name in ["Terminals", "Sheet1"]:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        terminals = []
        for row in rows:
            if not row or not row[0]:
                continue
            terminals.append(Terminal(
                name=str(row[0]).strip(),
                city=str(row[1]).strip() if len(row) > 1 and row[1] else "",
                state=str(row[2]).strip() if len(row) > 2 and row[2] else "",
                delivery_method=str(row[3]).strip() if len(row) > 3 and row[3] else "",
                supplier=str(row[4]).strip() if len(row) > 4 and row[4] else "",
            ))

        wb.close()
        logger.info(f"Loaded {len(terminals)} terminals")
        return terminals

    def load_transport_rates(self, path: Optional[Path] = None) -> list[TransportRate]:
        """Load transport rate data from Terminals.xlsx FBOs sheet."""
        try:
            import openpyxl
        except ImportError:
            return []

        xlsx_path = path or self._find_file("Terminals.xlsx")
        if not xlsx_path:
            return []

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "FBOs" not in wb.sheetnames:
            wb.close()
            return []

        ws = wb["FBOs"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rates = []
        for row in rows:
            if not row or not row[0]:
                continue
            fbo = str(row[0]).strip()
            terminal = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            miles = _try_float(row[5]) if len(row) > 5 else None
            rate = _try_float(row[10]) if len(row) > 10 else None

            if terminal and rate is not None:
                rates.append(TransportRate(
                    origin_terminal=terminal,
                    destination_fbo=fbo,
                    distance_miles=miles or 0.0,
                    rate_per_gal=rate,
                ))

        wb.close()
        logger.info(f"Loaded {len(rates)} transport rates")
        return rates

    def load_platts_sample(self, path: Optional[Path] = None) -> list[dict]:
        """Load the Gulf Coast Platts sample data for backtesting."""
        try:
            import openpyxl
        except ImportError:
            return []

        xlsx_path = path or self._find_file(
            "GC_Platts_Weekday_Sample_2023-01-02_to_2025-08-08.xlsx"
        )
        if not xlsx_path:
            return []

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        records = []
        for row in rows:
            if not row or not row[0]:
                continue
            dt = row[0]
            if hasattr(dt, "strftime"):
                date_str = dt.strftime("%Y-%m-%d")
            else:
                date_str = str(dt)[:10]

            pda = _try_float(row[2])
            pwa = _try_float(row[3])
            if pda is not None:
                records.append({
                    "date": date_str,
                    "prior_day": pda,
                    "prior_week_avg": pwa,
                })

        wb.close()
        logger.info(f"Loaded {len(records)} Platts sample records")
        return records

    def export_fbos_json(self, fbos: list[FBO], output_path: Optional[Path] = None):
        """Export FBO data as JSON for quick loading without openpyxl."""
        path = output_path or self.ref_dir / "fbos.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        data = []
        for f in fbos:
            data.append({
                "code": f.code,
                "city": f.city,
                "state": f.state,
                "latitude": f.latitude,
                "longitude": f.longitude,
                "padd": f.padd,
                "index_publisher": f.index_publisher,
                "index_name": f.index_name,
                "sample_differential": f.sample_differential,
                "region_label": f.region_label,
            })
        path.write_text(json.dumps(data, indent=2))
        logger.info(f"Exported {len(data)} FBOs to {path}")

    def _find_file(self, *names: str) -> Optional[Path]:
        """Search for a file in reference dir, then Downloads."""
        for name in names:
            for search_dir in [self.ref_dir, DOWNLOADS]:
                p = search_dir / name
                if p.exists():
                    return p
        return None
